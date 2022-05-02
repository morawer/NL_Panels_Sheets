import os
import tkinter as tk
from tkinter import filedialog

import openpyxl
import pandas as pd
from reportlab.pdfgen import canvas

#Sheets sizes
width = 841
height = 595
A4 = [width, height]

#Declare variables
moNew = 0
newWeek = False
weekNumber = 0
rowTotal = 0
sheetAcum = 0
panelsAcum = 0

#Function to select the excel we want to extract the data.
def fileSelection():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()
    return file_path

def msg_error(co, line, mo):
    print("===========================================================================")
    print("         |>>> ERROR AL CREAR ==> CO: " + str(co[line]) + "_ MO:" + str(mo[line]) + " <<<|  ")
    print("===========================================================================")

#Destination folder of PDF sheets and Excel files created
pathDestination = "/home/dani/Projects/NL_Panels_Sheets/destination/"

#If destination folder does not exist, the folder is created automatically
if not os.path.exists(pathDestination):
    os.makedirs(pathDestination)
    os.makedirs(pathDestination + 'etiquetas/')

#The info input is trough of excel file
excel = fileSelection()
print(excel)
df = pd.read_excel(excel, sheet_name="Database")

#In "Overview panels ESP 0426.xlsx" file we get in variables the differents columns values we need.
mo = df["Ref order no"].values
co = df["CO España"].values
itemNumber = df["Item no"].values
itemName = df["Item name"].values
po = df["Order no PO"].values
date = df["New Date ES"].values
qty = df["Ordered qty"].values
week = df["Week"].values
lastRow = len(df.index)
for line in range(len(mo)):
    if mo[line] != moNew:
        try:
            pdf_panelsNL = canvas.Canvas(
                pathDestination + str(mo[line]) + "_" + "PanelsNL" + ".pdf", pagesize=A4)
            pdf_panelsNL.setFont('Helvetica-Bold', 72)
            pdf_panelsNL.drawCentredString(
                width/2, height - 100, 'PAN-PUE-INT')
            pdf_panelsNL.drawCentredString(
                width/2, height - 220, "PO: " + str(po[line]))
            pdf_panelsNL.drawCentredString(width/2, height - 340, "MO: " + str(mo[line]))
            pdf_panelsNL.drawCentredString(width/2, height - 460, 'HOLANDA')
            pdf_panelsNL.save()
            print(str(mo[line]) + "_" + "PanelsNL" + ".pdf >>>>> SHEET CREATED")
            moNew = mo[line]
            sheetAcum = sheetAcum + 1
        except:
            msg_error(co, line, mo)
            
print(f'>>> It have been created {sheetAcum} sheets. <<<')
print('*******************************************')

for line in range(len(co)):
    if weekNumber != week[line]:
        if weekNumber != 0:
            excel_labels.save(f'{pathDestination}etiquetas/ETIQUETAS SEMANA {weekNumber}.xlsx')
            rowTotal = 0  
            print(f'>>> WEEK {weekNumber} >>> {panelsAcum} panels')
            panelsAcum = 0
        #Create the excel
        excel_labels = openpyxl.Workbook()
        sheet_labels = excel_labels.active
        #Write the title in every column
        cell_A1 = sheet_labels.cell(row=1, column=1)
        cell_A1.value = 'Item Nr'
        cell_A2 = sheet_labels.cell(row=1, column=2)
        cell_A2.value = 'Descripción'
        cell_A3 = sheet_labels.cell(row=1, column=3)
        cell_A3.value = 'MO'
        cell_A4 = sheet_labels.cell(row=1, column=4)
        cell_A4.value = 'PO'
        cell_A5 = sheet_labels.cell(row=1, column=5)
        cell_A5.value = 'CO España'
        weekNumber = week[line]
        
    if weekNumber == week[line]:
        for item in range (0, qty[line]):
            cell_itemNumber = sheet_labels.cell(row=2 + rowTotal, column=1)
            cell_itemNumber.value = itemNumber[line]
            cell_description = sheet_labels.cell(row=2 + rowTotal, column=2)
            cell_description.value = itemName[line]
            cell_mo = sheet_labels.cell(row=2 + rowTotal, column=3)
            cell_mo.value = mo[line]
            cell_po = sheet_labels.cell(row=2 + rowTotal, column=4)
            cell_po.value = po[line]
            cell_co = sheet_labels.cell(row=2 + rowTotal, column=5)
            cell_co.value = co[line]
            rowTotal = rowTotal + 1
            panelsAcum = panelsAcum + 1

    if line == lastRow-1:
        excel_labels.save(f'{pathDestination}etiquetas/ETIQUETAS SEMANA {weekNumber}.xlsx')
        print(f'>>> WEEK {weekNumber} >>> {panelsAcum} panels')
        panelsAcum = 0
        
            